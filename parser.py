import openpyxl
import connector
import os


def ParsDateFromExel(sheet):
    DataRasp = []
    Max_row = sheet.max_row
    if sheet.max_column == 4:
        num_row = 'D' + str(Max_row)
        for row in sheet['A2': num_row]:
            string = ''
            for cell in row:
                string = string + str(cell.value) + '   '
            DataRow = string.split('   ')
            DataRasp.append(DataRow)

        for i in range(len(DataRasp)):
            # Удаляем последний пустой символ
            del DataRasp[i][-1]
            # Выставляем дни недели в каждую строчку
            if DataRasp[i][0] == "None":
                DataRasp[i][0] = DataRasp[i - 1][0]
        # Строки без пар
        for i in range(len(DataRasp)):
            if DataRasp[i][3] == "None":
                DataRasp[i].clear()
        return DataRasp
    else:
        num_row = 'F' + str(Max_row)
        for row in sheet['A2': num_row]:
            string = ''
            for cell in row:
                string = string + str(cell.value) + '   '
            DataRow = string.split('   ')
            DataRasp.append(DataRow)

        for i in range(len(DataRasp)):
            # Удаляем последний пустой символ
            del DataRasp[i][-1]
            # Выставляем дни недели в каждую строчку
            if DataRasp[i][0] == "None":
                DataRasp[i][0] = DataRasp[i - 1][0]
        # Строки без пар
        for i in range(len(DataRasp)):
            if DataRasp[i][3] == "None":
                DataRasp[i].clear()
        return DataRasp


def GroupName(sheet):
    if sheet.max_column == 4:
        group = sheet['C1'].value
        return group
    else:
        group1 = sheet['C1'].value
        group2 = sheet['E1'].value
        return group1, group2


def AddToBase(Value):
    command = 'INSERT INTO parsData (groupName , day, numLesson, dataLesson, NumberClass) VALUES (?, ?, ?, ?, ?)'
    connector.ValuesExecuteSQL(command, Value)
    connector.Commit()


def InsertBase(sheet):
    Rasp = ParsDateFromExel(sheet)
    for i in range(len(Rasp)):
        if len(Rasp[i]) != 0:
            if len(Rasp[i]) == 4:
                gr = GroupName(sheet)
                day = Rasp[i][0]
                numLess = Rasp[i][1]
                data = Rasp[i][2]
                cl = Rasp[i][3]
                Value = [gr, day, numLess, data, cl]
                AddToBase(Value)
            else:
                gr = GroupName(sheet)
                gr1 = gr[0]
                gr2 = gr[1]
                day = Rasp[i][0]
                numLess = Rasp[i][1]
                data1 = Rasp[i][2]
                cl1 = Rasp[i][3]
                data2 = Rasp[i][4]
                cl2 = Rasp[i][5]
                Value1 = [gr1, day, numLess, data1, cl1]
                Value2 = [gr2, day, numLess, data2, cl2]
                AddToBase(Value1)
                AddToBase(Value2)
        else:
            continue


def parsFile():
    rod = os.getcwd()
    directory = rod + '/xlsx/'
    list_file = os.listdir(directory)
    files = []
    for i in range(len(list_file)):
        tempFile = directory + list_file[i]
        if ".xlsx" in tempFile:
            files.append(tempFile)
        else:
            continue
    for i in range(len(files)):
        book = openpyxl.load_workbook(files[i])
        sheet = book.active
        InsertBase(sheet)
        book.close()
        os.remove(files[i])



